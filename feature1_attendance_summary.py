"""
Feature 1: 근태 요약 텍스트 생성
명세: feature-1-spec.md

입력: 일일근태현황 엑셀 파일 (시트 HuDut020ControlDB, 2행 헤더, 3행부터 데이터)
출력: 팀장근태/팀원근태로 구분된, 근태 종류별 인원수(소속팀 포함) 요약 텍스트 1개
"""

import sys
from datetime import datetime

import openpyxl

SHEET_NAME = "HuDut020ControlDB"
WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]


def load_rows(xlsx_path):
    """1) 엑셀 파일을 열어 지정 시트를 읽는다 (2행 헤더, 3행부터 데이터)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    header = [cell.value for cell in ws[2]]
    col_index = {name: i for i, name in enumerate(header)}

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        rows.append(row)
    return rows, col_index


def select_columns(rows, col_index):
    """2) 필요한 열만 남긴다: 근태일(B, 중복 판단용), 성명(F), 부서명(H), 근태코드(I), 비고(M)."""
    selected = []
    for row in rows:
        selected.append({
            "근태일": row[col_index["근태일"]],
            "성명": row[col_index["성명"]],
            "부서명": row[col_index["부서명"]],
            "근태코드": row[col_index["근태코드"]],
            "비고": row[col_index["비고"]],
        })
    return selected


def remove_blank_rows(records):
    """3) 빈 행을 제거한다."""
    return [r for r in records if any(v not in (None, "") for v in r.values())]


def remove_duplicate_rows(records):
    """4) 중복 행을 제거한다."""
    seen = set()
    result = []
    for r in records:
        key = tuple(r.items())
        if key not in seen:
            seen.add(key)
            result.append(r)
    return result


def split_leader_member(records):
    """5) H열(부서명)이 '팀장'으로 끝나면 팀장근태, 그 외는 팀원근태로 분리."""
    leader, member = [], []
    for r in records:
        dept = r["부서명"] or ""
        if dept.endswith("팀장"):
            leader.append(r)
        else:
            member.append(r)
    return leader, member


def clean_dept_name(dept):
    """이미 다 아는 정보인 '(발안)'/'-발안' 표기는 부서명에서 생략한다."""
    return (dept or "").replace("(발안)", "").replace("-발안", "")


def build_summary_phrase(records, show_team_count=True):
    """6~9) 근태 종류별 그룹핑 -> 인원수 집계 -> 소속팀 세부 정리 -> 요약 문구 생성.

    show_team_count=False면 팀명 뒤 인원수를 붙이지 않는다.
    (팀장근태는 팀장 부서마다 1명뿐이라 숫자 표기가 불필요)
    """
    # 근태코드 등장 순서를 유지하기 위해 순서를 기록
    order = []
    by_code = {}
    for r in records:
        code = r["근태코드"]
        if code not in by_code:
            by_code[code] = []
            order.append(code)
        by_code[code].append(r)

    phrases = []
    for code in order:
        items = by_code[code]
        total = len(items)

        team_order = []
        team_count = {}
        for r in items:
            dept = clean_dept_name(r["부서명"])
            if dept not in team_count:
                team_count[dept] = 0
                team_order.append(dept)
            team_count[dept] += 1

        if show_team_count:
            team_parts = ", ".join(f"{dept}{team_count[dept]}" for dept in team_order)
        else:
            team_parts = ", ".join(team_order)
        phrases.append(f"{code} {total}명({team_parts})")

    return phrases


def merge_summaries(leader_phrases, member_phrases):
    """10) 팀장근태 요약 + 팀원근태 요약 병합."""
    return {
        "팀장근태": leader_phrases,
        "팀원근태": member_phrases,
    }


def resolve_report_date(records):
    """데이터의 실제 근태일(B열)에서 보고서 날짜를 뽑는다. 못 찾으면 오늘 날짜로 대체."""
    for r in records:
        raw = r.get("근태일")
        if not raw:
            continue
        if isinstance(raw, datetime):
            return raw
        text = str(raw).strip().split(" ")[0]  # "2026-07-23 00:00:00.0" -> "2026-07-23"
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return datetime.now()


def build_title(report_date):
    """근태일·요일 기반 제목을 만든다 (예: 26년7월23일(목) 발안공장 일일근태보고)."""
    weekday = WEEKDAY_KO[report_date.weekday()]
    return f"{report_date.strftime('%y')}년{report_date.month}월{report_date.day}일({weekday}) 발안공장 일일근태보고"


def build_total_summary(records):
    """전체 근태 인원을 근태코드별로 묶어 '1. 근태요약' 블록을 만든다 (원래 근태코드 그대로 표기, 4개씩 줄바꿈)."""
    order = []
    by_code = {}
    for r in records:
        code = r["근태코드"]
        if code not in by_code:
            by_code[code] = 0
            order.append(code)
        by_code[code] += 1

    total = sum(by_code.values())
    items = [f"{code}{by_code[code]}" for code in order]

    grouped_lines = [", ".join(items[i:i + 4]) for i in range(0, len(items), 4)]

    lines = [f"1. 근태요약 : 총 {total}명"]
    for i, line in enumerate(grouped_lines):
        prefix = "(" if i == 0 else " "
        suffix = ")" if i == len(grouped_lines) - 1 else ""
        lines.append(f"{prefix}{line}{suffix}")

    return lines


def format_for_kakao(merged, total_lines, report_date):
    """11) 카톡 발송용 포맷으로 변환. 근태 종류가 많으면 줄을 나눠서 표기한다."""
    lines = []
    lines.append(build_title(report_date))
    lines.append("")
    lines.extend(total_lines)
    lines.append("")
    lines.append("2. 팀장근태")
    lines.extend(merged["팀장근태"] if merged["팀장근태"] else ["해당 없음"])
    lines.append("")
    lines.append("3. 팀원근태")
    lines.extend(merged["팀원근태"] if merged["팀원근태"] else ["해당 없음"])
    return "\n".join(lines)


def generate_attendance_summary(xlsx_path):
    rows, col_index = load_rows(xlsx_path)
    records = select_columns(rows, col_index)
    records = remove_blank_rows(records)
    records = remove_duplicate_rows(records)
    leader_records, member_records = split_leader_member(records)
    leader_phrase = build_summary_phrase(leader_records, show_team_count=False)
    member_phrase = build_summary_phrase(member_records)
    merged = merge_summaries(leader_phrase, member_phrase)
    total_lines = build_total_summary(records)
    report_date = resolve_report_date(records)
    return format_for_kakao(merged, total_lines, report_date)


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "inputs/sample-attendance-fake.xlsx"
    print(generate_attendance_summary(xlsx_path))
